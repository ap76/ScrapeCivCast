from bs4 import BeautifulSoup
from selenium import webdriver
import time
import openpyxl

driver = webdriver.Firefox()
driver.implicitly_wait(30)

wb = openpyxl.Workbook()
row_excel = 0

with open("list", "r") as list_file:
    for url in list_file:
        row_excel += 1
        print(row_excel)
        print(url)
        driver.get(url)
        time.sleep(3)
        outerHTML = driver.execute_script("return document.documentElement.outerHTML")
        html_soup = BeautifulSoup(outerHTML, 'html.parser')
        #url sent to the Excel file
        wb.active.cell(row=row_excel, column=1).value = str(url)

        # Project scope sent to the Excel file
        project_scope = []
        for i in html_soup.find_all('div', class_='card-body'):
            project_scope.append(i.text)
        # for i in range(len(project_scope)):
        # print(project_scope[4])
        if len(project_scope)>1:
            wb.active.cell(row=row_excel, column=2).value = str(project_scope[4]).strip()

        # Project description sent to the Excel file
        cc = []
        for i in html_soup.find_all('div', class_='col-8 col-md-7 col-xl-8'):
            # print (i.text)
            cc.append(i.text)
        for i in range(len(cc)):
            # print(cc[i])
            wb.active.cell(row=row_excel, column=i + 3).value = str(cc[i]).strip()
        wb.save('scraped.xlsx')
wb.close()
